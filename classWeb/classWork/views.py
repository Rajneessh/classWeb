from re import X
from sre_compile import isstring
from fastnumbers import isfloat
from django.shortcuts import render
import openpyxl
from requests import delete
# Create your views here.
def classHome(request):
    return render(request, 'classWork/classHome.html')

def timeTable(request):
    return render(request, 'classWork/timetable.html')


def home(request):
    return render(request,'classWork/home.html')

def excel(request):

    # First time loading the page
    if "GET" == request.method:
        return render(request, 'classWork/ExcelProcessing.html')
    
    # Uploaded an excel file
    else:
        excelFile = request.FILES['File']    
        wb = openpyxl.load_workbook(excelFile) 
        
        sheets = wb.sheetnames
        print(sheets)
        
        for i in range(len(sheets)):
            worksheet = wb[sheets[i]]
            print(worksheet)
            
            active_sheet = wb.active
            # print(active_sheet)
            
            # print(worksheet['A1'].value)
            
            excelData = list()
            
            sum = 0
            numList = []
            floatList = []
            stringList = []
            listList = []
            for row in worksheet.iter_rows():
                rowData = list()
                for cell in row:
                    cell.value
                    if None == cell.value:
                        continue
                    if cell.value is not None:
                        rowData.append(str(cell.value))
                        # print(cell.value)
                    
                    
                    if str(cell.value).isdigit() == True:
                        numList.append(cell.value)
                        sum = sum + cell.value
                    elif isfloat(str(cell.value)) == True:
                        floatList.append(cell.value)
                        sum = sum+ cell.value
                        
                    else:
                        if cell.value == '':
                            continue
                        else:
                            stringList.append(str(cell.value))

                    
                
                excelData.append(rowData)
            
            
                
            return render(request, 'classWork/ExcelProcessing.html',
                        {'excelData':excelData, 'sum':sum,
                        'numList':numList,'floatList':floatList,
                        'stringList':stringList, 'listList':listList})